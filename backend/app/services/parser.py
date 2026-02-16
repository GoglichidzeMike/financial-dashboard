from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import io
import re
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = ("date", "details", "gel", "usd", "eur", "gbp")
CURRENCY_HEADERS = ("gel", "usd", "eur", "gbp")

_AMOUNT_RE = re.compile(
    r"Amount\s*:?\s*(?P<currency>[A-Z]{3})\s*(?P<amount>[-+]?\d[\d\s\u00a0.,]*)",
    re.IGNORECASE,
)
_RATE_RE = re.compile(r"rate\s*:\s*(?P<rate>\d+(?:[.,]\d+)?)", re.IGNORECASE)
_MCC_RE = re.compile(r"MCC\s*:\s*(?P<mcc>\d+)", re.IGNORECASE)
_CARD_RE = re.compile(r"Card\s*No\s*:\s*\*+(?P<last4>\d{4})", re.IGNORECASE)
_POSTED_DATE_RE = re.compile(
    r"Date\s*:\s*(?P<d>\d{2}/\d{2}/\d{4})(?:\s+\d{2}:\d{2})?", re.IGNORECASE
)


class ParserError(ValueError):
    pass


@dataclass(slots=True)
class ParsedTransaction:
    date: date
    posted_date: date | None
    description_raw: str
    direction: str
    amount_original: Decimal
    currency_original: str
    amount_gel: Decimal
    conversion_rate: Decimal | None
    card_last4: str | None
    mcc_code: str | None
    dedup_key: str


@dataclass(slots=True)
class ParseResult:
    transactions: list[ParsedTransaction]
    rows_total: int
    rows_skipped_non_transaction: int
    rows_invalid: int



def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace('"', " ").replace("\n", " ").strip().lower()
    return re.sub(r"\s+", " ", text)



def _normalize_description_for_hash(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())



def compute_dedup_key(txn_date: date, amount_original: Decimal, description_raw: str) -> str:
    description = _normalize_description_for_hash(description_raw)
    canonical = f"{txn_date.isoformat()}|{amount_original.quantize(Decimal('0.01'))}|{description}"
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()



def parse_decimal_value(value: Any) -> Decimal:
    if value is None:
        raise InvalidOperation("empty")

    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        raise InvalidOperation("blank")

    text = text.replace("\u00a0", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    return Decimal(text)



def infer_direction(details: str) -> str:
    lowered = details.strip().lower()
    if lowered.startswith("payment"):
        return "expense"
    if lowered.startswith("income"):
        return "income"
    if lowered.startswith("incoming transfer") or "transfer" in lowered:
        return "transfer"
    return "expense"



def _parse_statement_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    return datetime.strptime(text, "%d/%m/%Y").date()



def _parse_posted_date(details: str) -> date | None:
    match = _POSTED_DATE_RE.search(details)
    if not match:
        return None
    return datetime.strptime(match.group("d"), "%d/%m/%Y").date()



def _parse_amount_from_details(details: str) -> tuple[str, Decimal] | None:
    match = _AMOUNT_RE.search(details)
    if not match:
        return None

    currency = match.group("currency").upper()
    amount = abs(parse_decimal_value(match.group("amount")))
    return currency, amount



def _parse_conversion_rate(details: str) -> Decimal | None:
    match = _RATE_RE.search(details)
    if not match:
        return None
    return parse_decimal_value(match.group("rate"))



def _find_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    max_scan_rows = min(ws.max_row, 40)
    for row_idx in range(1, max_scan_rows + 1):
        row = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        normalized = [_normalize_header(cell) for cell in row]

        header_map: dict[str, int] = {}
        for col_idx, cell in enumerate(normalized):
            if cell == "date":
                header_map["date"] = col_idx
            elif cell == "details":
                header_map["details"] = col_idx
            elif cell in CURRENCY_HEADERS:
                header_map[cell] = col_idx

        if all(key in header_map for key in REQUIRED_HEADERS):
            return row_idx, header_map

    raise ParserError("Could not find required headers: Date, Details, GEL, USD, EUR, GBP")



def _extract_signed_currency_value(row_values: dict[str, Any]) -> tuple[str | None, Decimal | None]:
    for currency in CURRENCY_HEADERS:
        value = row_values.get(currency)
        if value is None or str(value).strip() == "":
            continue
        try:
            return currency.upper(), parse_decimal_value(value)
        except InvalidOperation:
            continue
    return None, None



def parse_statement_xlsx(file_bytes: bytes) -> ParseResult:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParserError(f"Failed to read XLSX file: {exc}") from exc

    parsed: list[ParsedTransaction] = []
    rows_total = 0
    rows_skipped = 0
    rows_invalid = 0

    worksheet_and_header: tuple[Any, int, dict[str, int]] | None = None
    for ws in workbook.worksheets:
        try:
            header_row_idx, header_map = _find_header_row(ws)
            worksheet_and_header = (ws, header_row_idx, header_map)
            break
        except ParserError:
            continue

    if worksheet_and_header is None:
        raise ParserError("Could not find a worksheet with required statement headers")

    ws, header_row_idx, header_map = worksheet_and_header

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_values = {
            key: ws.cell(row=row_idx, column=col_idx + 1).value
            for key, col_idx in header_map.items()
        }

        if all(v is None or str(v).strip() == "" for v in row_values.values()):
            continue

        rows_total += 1

        date_cell = row_values.get("date")
        details_cell = row_values.get("details")
        details = "" if details_cell is None else str(details_cell).strip()

        if isinstance(date_cell, str) and date_cell.strip().lower() == "balance":
            rows_skipped += 1
            continue

        normalized_date_cell = _normalize_header(date_cell)
        normalized_details_cell = _normalize_header(details_cell)
        if normalized_date_cell == "date" and normalized_details_cell == "details":
            rows_skipped += 1
            continue

        if not details and all(
            row_values.get(currency) in (None, "") for currency in CURRENCY_HEADERS
        ):
            rows_skipped += 1
            continue

        try:
            statement_date = _parse_statement_date(date_cell)
            posted_date = _parse_posted_date(details)
            direction = infer_direction(details)

            details_amount = _parse_amount_from_details(details)
            table_currency, table_signed_amount = _extract_signed_currency_value(row_values)
            if details_amount:
                currency_original = details_amount[0]
                amount_original = details_amount[1]
            else:
                if table_currency is None or table_signed_amount is None:
                    raise ValueError("Missing amount information")
                currency_original = table_currency
                amount_original = abs(table_signed_amount)

            conversion_rate = _parse_conversion_rate(details)

            gel_cell = row_values.get("gel")
            amount_gel: Decimal
            if gel_cell is not None and str(gel_cell).strip() != "":
                amount_gel = abs(parse_decimal_value(gel_cell))
            elif currency_original != "GEL" and conversion_rate is not None:
                amount_gel = (amount_original * conversion_rate).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            elif currency_original == "GEL":
                amount_gel = amount_original
            elif table_signed_amount is not None:
                amount_gel = abs(table_signed_amount)
            else:
                raise ValueError("Unable to derive GEL amount")

            mcc_match = _MCC_RE.search(details)
            card_match = _CARD_RE.search(details)
            mcc_code = mcc_match.group("mcc") if mcc_match else None
            card_last4 = card_match.group("last4") if card_match else None

            parsed.append(
                ParsedTransaction(
                    date=statement_date,
                    posted_date=posted_date,
                    description_raw=details,
                    direction=direction,
                    amount_original=amount_original.quantize(Decimal("0.01")),
                    currency_original=currency_original,
                    amount_gel=amount_gel.quantize(Decimal("0.01")),
                    conversion_rate=conversion_rate,
                    card_last4=card_last4,
                    mcc_code=mcc_code,
                    dedup_key=compute_dedup_key(statement_date, amount_original, details),
                )
            )
        except Exception:  # noqa: BLE001
            rows_invalid += 1

    return ParseResult(
        transactions=parsed,
        rows_total=rows_total,
        rows_skipped_non_transaction=rows_skipped,
        rows_invalid=rows_invalid,
    )
